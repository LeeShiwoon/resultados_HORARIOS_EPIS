"""
Módulo de exportación de horarios a PDF y Excel con subcolunas dinámicas
"""
import hashlib
from pathlib import Path
from reportlab.lib.pagesizes import landscape, A4
from reportlab.lib.units import cm
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, PageBreak, Spacer
from reportlab.lib import colors
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from horario import DIAS

OUTPUT_DIR = Path("output")
OUTPUT_DIR.mkdir(exist_ok=True)

def generar_color_pastel(texto):
    hash_obj = hashlib.md5(texto.encode())
    hash_hex = hash_obj.hexdigest()
    r = min(220, max(140, int(int(hash_hex[0:2], 16) * 0.5 + 100)))
    g = min(220, max(140, int(int(hash_hex[2:4], 16) * 0.5 + 100)))
    b = min(220, max(140, int(int(hash_hex[4:6], 16) * 0.5 + 100)))
    return f"{r:02x}{g:02x}{b:02x}"

def exportar_a_pdf(df_horario, max_colisiones, franjas_activas, ciclo):
    file_path = OUTPUT_DIR / f"Horario_Ciclo_{ciclo}.pdf"
    doc = SimpleDocTemplate(str(file_path), pagesize=landscape(A4),
                           rightMargin=0.3*cm, leftMargin=0.3*cm,
                           topMargin=0.5*cm, bottomMargin=0.3*cm)
    elements = []
    styles = getSampleStyleSheet()
    
    title = Paragraph(f"CICLO {ciclo} - Horario Académico", 
                     ParagraphStyle('CustomTitle', parent=styles['Heading1'],
                                   fontSize=14, textColor=colors.HexColor('#2F5496'),
                                   spaceAfter=4, alignment=1, fontName='Helvetica-Bold'))
    elements.append(title)
    elements.append(Spacer(1, 0.2*cm))
    
    datos_tabla = []
    encabezado = ["HORARIO"]
    for dia in DIAS:
        num_subcols = max_colisiones[dia]
        encabezado.append(dia)
        for _ in range(1, num_subcols):
            encabezado.append("")
    datos_tabla.append(encabezado)
    
    cell_style = ParagraphStyle('CellStyle', parent=styles['Normal'],
                               fontSize=6, alignment=1, wordWrap='CJK',
                               fontName='Times-Roman', leading=7)
    
    for franja in franjas_activas:
        fila = [Paragraph(franja, ParagraphStyle('FranjaStyle', parent=styles['Normal'],
                                                fontSize=6, alignment=1,
                                                fontName='Times-Bold', leading=6))]
        for dia in DIAS:
            cursos = df_horario[dia][franja]
            num_subcols = max_colisiones[dia]
            while len(cursos) < num_subcols:
                cursos.append("")
            for idx_sub in range(num_subcols):
                valor = cursos[idx_sub] if idx_sub < len(cursos) else ""
                fila.append(Paragraph(valor if valor else "", cell_style))
        datos_tabla.append(fila)
    
    page_width = landscape(A4)[0]
    ancho_horario = 1.5*cm
    ancho_total = page_width - 0.6*cm
    total_subcols = sum(max_colisiones.values())
    ancho_subcolumna = (ancho_total - ancho_horario) / total_subcols
    col_widths = [ancho_horario] + [ancho_subcolumna] * total_subcols
    
    tabla = Table(datos_tabla, colWidths=col_widths, repeatRows=1)
    
    comandos = [
        ('FONTNAME', (0, 0), (-1, -1), 'Times-Roman'),
        ('FONTSIZE', (0, 0), (-1, -1), 6),
        ('ALIGNMENT', (0, 0), (-1, -1), 'CENTER'),
        ('VALIGNMENT', (0, 0), (-1, -1), 'TOP'),
        ('TOPPADDING', (0, 0), (-1, -1), 2),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 2),
        ('LEFTPADDING', (0, 0), (-1, -1), 2),
        ('RIGHTPADDING', (0, 0), (-1, -1), 2),
        ('FONTNAME', (0, 0), (-1, 0), 'Times-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 7),
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#2F5496')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('BACKGROUND', (0, 1), (0, -1), colors.HexColor('#D0CECE')),
        ('FONTNAME', (0, 1), (0, -1), 'Times-Bold'),
        ('FONTSIZE', (0, 1), (0, -1), 6),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
    ]
    
    col_actual = 1
    for dia in DIAS:
        num_subcols = max_colisiones[dia]
        if num_subcols > 1:
            comandos.append(('SPAN', (col_actual, 0), (col_actual + num_subcols - 1, 0)))
        col_actual += num_subcols
    
    for i in range(1, len(datos_tabla)):
        for j in range(1, len(datos_tabla[i])):
            valor_obj = datos_tabla[i][j]
            valor_text = valor_obj.text if hasattr(valor_obj, 'text') else str(valor_obj)
            if valor_text and valor_text.strip():
                asignatura = valor_text.split(' - ')[0].strip()
                comandos.append(('BACKGROUND', (j, i), (j, i), 
                               colors.HexColor(f"#{generar_color_pastel(asignatura)}")))
    
    tabla.setStyle(TableStyle(comandos))
    elements.append(tabla)
    doc.build(elements)
    print(f"📄 PDF exportado: {file_path}")
    return str(file_path)

def exportar_todos_pdf(horarios_dict, max_colisiones_dict, franjas_activas_dict):
    file_path = OUTPUT_DIR / "Horarios_Completo.pdf"
    doc = SimpleDocTemplate(str(file_path), pagesize=landscape(A4),
                           rightMargin=0.3*cm, leftMargin=0.3*cm,
                           topMargin=0.5*cm, bottomMargin=0.3*cm)
    elements = []
    styles = getSampleStyleSheet()
    ciclos = sorted(list(horarios_dict.keys()))
    
    for idx_ciclo, ciclo in enumerate(ciclos):
        df_horario = horarios_dict[ciclo]
        max_colisiones = max_colisiones_dict[ciclo]
        franjas_activas = franjas_activas_dict[ciclo]
        
        title = Paragraph(f"CICLO {ciclo} - Horario Académico",
                         ParagraphStyle('CustomTitle', parent=styles['Heading1'],
                                       fontSize=14, textColor=colors.HexColor('#2F5496'),
                                       spaceAfter=4, alignment=1, fontName='Helvetica-Bold'))
        elements.append(title)
        elements.append(Spacer(1, 0.2*cm))
        
        datos_tabla = []
        encabezado = ["HORARIO"]
        for dia in DIAS:
            num_subcols = max_colisiones[dia]
            encabezado.append(dia)
            for _ in range(1, num_subcols):
                encabezado.append("")
        datos_tabla.append(encabezado)
        
        cell_style = ParagraphStyle('CellStyle', parent=styles['Normal'],
                                   fontSize=6, alignment=1, wordWrap='CJK',
                                   fontName='Times-Roman', leading=7)
        
        for franja in franjas_activas:
            fila = [Paragraph(franja, ParagraphStyle('FranjaStyle', parent=styles['Normal'],
                                                    fontSize=6, alignment=1,
                                                    fontName='Times-Bold', leading=6))]
            for dia in DIAS:
                cursos = df_horario[dia][franja]
                num_subcols = max_colisiones[dia]
                while len(cursos) < num_subcols:
                    cursos.append("")
                for idx_sub in range(num_subcols):
                    valor = cursos[idx_sub] if idx_sub < len(cursos) else ""
                    fila.append(Paragraph(valor if valor else "", cell_style))
            datos_tabla.append(fila)
        
        page_width = landscape(A4)[0]
        ancho_horario = 1.5*cm
        ancho_total = page_width - 0.6*cm
        total_subcols = sum(max_colisiones.values())
        ancho_subcolumna = (ancho_total - ancho_horario) / total_subcols
        col_widths = [ancho_horario] + [ancho_subcolumna] * total_subcols
        
        tabla = Table(datos_tabla, colWidths=col_widths, repeatRows=1)
        
        comandos = [
            ('FONTNAME', (0, 0), (-1, -1), 'Times-Roman'),
            ('FONTSIZE', (0, 0), (-1, -1), 6),
            ('ALIGNMENT', (0, 0), (-1, -1), 'CENTER'),
            ('VALIGNMENT', (0, 0), (-1, -1), 'TOP'),
            ('TOPPADDING', (0, 0), (-1, -1), 2),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 2),
            ('LEFTPADDING', (0, 0), (-1, -1), 2),
            ('RIGHTPADDING', (0, 0), (-1, -1), 2),
            ('FONTNAME', (0, 0), (-1, 0), 'Times-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 7),
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#2F5496')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('BACKGROUND', (0, 1), (0, -1), colors.HexColor('#D0CECE')),
            ('FONTNAME', (0, 1), (0, -1), 'Times-Bold'),
            ('FONTSIZE', (0, 1), (0, -1), 6),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
        ]
        
        col_actual = 1
        for dia in DIAS:
            num_subcols = max_colisiones[dia]
            if num_subcols > 1:
                comandos.append(('SPAN', (col_actual, 0), (col_actual + num_subcols - 1, 0)))
            col_actual += num_subcols
        
        for i in range(1, len(datos_tabla)):
            for j in range(1, len(datos_tabla[i])):
                valor_obj = datos_tabla[i][j]
                valor_text = valor_obj.text if hasattr(valor_obj, 'text') else str(valor_obj)
                if valor_text and valor_text.strip():
                    asignatura = valor_text.split(' - ')[0].strip()
                    comandos.append(('BACKGROUND', (j, i), (j, i),
                                   colors.HexColor(f"#{generar_color_pastel(asignatura)}")))
        
        tabla.setStyle(TableStyle(comandos))
        elements.append(tabla)
        if idx_ciclo < len(ciclos) - 1:
            elements.append(PageBreak())
    
    doc.build(elements)
    print(f"📚 PDF completo: {file_path}")
    return str(file_path)

def exportar_horario_excel(df_horario, max_colisiones, franjas_activas, ciclo):
    """Exporta un horario individual a Excel con formato profesional."""
    file_path = OUTPUT_DIR / f"Horario_Ciclo_{ciclo}.xlsx"
    wb = Workbook()
    ws = wb.active
    ws.title = f"Ciclo {ciclo}"
    
    # Estilos
    thin_border = Border(
        left=Side(style='thin', color='000000'),
        right=Side(style='thin', color='000000'),
        top=Side(style='thin', color='000000'),
        bottom=Side(style='thin', color='000000')
    )
    
    header_fill = PatternFill(start_color='2F5496', end_color='2F5496', fill_type='solid')
    header_font = Font(bold=True, color='FFFFFF', size=11)
    time_fill = PatternFill(start_color='D0CECE', end_color='D0CECE', fill_type='solid')
    time_font = Font(bold=True, size=9)
    
    # Fila 1: Encabezados
    col_actual = 1
    ws.cell(row=1, column=col_actual, value="HORARIO")
    ws.cell(row=1, column=col_actual).fill = header_fill
    ws.cell(row=1, column=col_actual).font = header_font
    ws.cell(row=1, column=col_actual).alignment = Alignment(horizontal='center', vertical='center')
    ws.cell(row=1, column=col_actual).border = thin_border
    ws.column_dimensions[get_column_letter(col_actual)].width = 15
    col_actual += 1
    
    # Encabezados de días con merge
    for dia in DIAS:
        num_subcols = max_colisiones[dia]
        inicio_merge = col_actual
        fin_merge = col_actual + num_subcols - 1
        
        # Escribir en la primera celda
        ws.cell(row=1, column=inicio_merge, value=dia)
        
        # Si hay más de una subcolumna, hacer merge
        if num_subcols > 1:
            ws.merge_cells(start_row=1, start_column=inicio_merge, 
                          end_row=1, end_column=fin_merge)
        
        # Aplicar formato a todas las celdas del merge
        for col in range(inicio_merge, fin_merge + 1):
            cell = ws.cell(row=1, column=col)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = thin_border
            ws.column_dimensions[get_column_letter(col)].width = 45
        
        col_actual = fin_merge + 1
    
    # Filas de datos
    fila_actual = 2
    colores_asignaturas = {}
    
    for franja in franjas_activas:
        # Columna de franja horaria
        cell = ws.cell(row=fila_actual, column=1, value=franja)
        cell.fill = time_fill
        cell.font = time_font
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = thin_border
        
        # Celdas de clases
        col_actual = 2
        for dia in DIAS:
            cursos = df_horario[dia][franja]
            num_subcols = max_colisiones[dia]
            
            while len(cursos) < num_subcols:
                cursos.append("")
            
            for idx_sub in range(num_subcols):
                valor = cursos[idx_sub] if idx_sub < len(cursos) else ""
                cell = ws.cell(row=fila_actual, column=col_actual, value=valor)
                cell.alignment = Alignment(wrap_text=True, vertical='center', horizontal='center')
                cell.border = thin_border
                
                if valor and valor.strip():
                    asignatura = valor.split(' - ')[0].strip()
                    if asignatura not in colores_asignaturas:
                        colores_asignaturas[asignatura] = generar_color_pastel(asignatura)
                    color = colores_asignaturas[asignatura]
                    cell.fill = PatternFill(start_color=color, end_color=color, fill_type='solid')
                    cell.font = Font(size=9)
                else:
                    cell.fill = PatternFill(start_color='FFFFFF', end_color='FFFFFF', fill_type='solid')
                
                col_actual += 1
        
        ws.row_dimensions[fila_actual].height = 60
        fila_actual += 1
    
    wb.save(str(file_path))
    print(f"📊 Excel exportado: {file_path}")
    return str(file_path)

def exportar_todos_excel(horarios_dict, max_colisiones_dict, franjas_activas_dict):
    """Exporta todos los ciclos, cada uno en su propio archivo Excel."""
    archivos_generados = []
    ciclos = sorted(list(horarios_dict.keys()))
    
    for ciclo in ciclos:
        file_path = exportar_horario_excel(
            horarios_dict[ciclo],
            max_colisiones_dict[ciclo],
            franjas_activas_dict[ciclo],
            ciclo
        )
        archivos_generados.append(file_path)
    
    print(f"\n✅ Total de {len(archivos_generados)} archivos Excel generados")
    return archivos_generados